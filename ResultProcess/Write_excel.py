from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference, BarChart3D
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW


class Write_excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write(self, coord, value):
        # eg: coord:A1
        self.ws.cell(coord).value = value
        self.wb.save(self.filename)

    def merge(self, rangstring):
        # eg: rangstring:A1:E1
        self.ws.merge_cells(rangstring)
        self.wb.save(self.filename)

    def cellstyle(self, coord, font, align):
        cell = self.ws.cell(coord)
        cell.font = font
        cell.alignment = align

    def makechart(self, title, pos, width, height, col1, row1, col2, row2, col3, row3, row4):
        ''''':param title:图表名
                  pos:图表位置
                  width:图表宽度
                  height:图表高度
        '''
        data = Reference(self.ws, min_col=col1, min_row=row1, max_col=col2, max_row=row2)
        cat = Reference(self.ws, min_col=col3, min_row=row3, max_row=row4)
        chart = BarChart3D()
        chart.title = title
        chart.width = width
        chart.height = height
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(cat)
        self.ws.add_chart(chart, pos)
        self.wb.save(self.filename)

