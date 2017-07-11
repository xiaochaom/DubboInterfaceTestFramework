#coding:utf-8
"""
Author:李亚超
Date:20161027
"""
import sys, os,logging

reload(sys)
sys.setdefaultencoding('utf8')
sys.path.append("..")

import openpyxl
from Config.EncodeConf import MyEncode
from openpyxl.styles import Font, colors, Border, Side,PatternFill,Alignment

from openpyxl.workbook import Workbook


reload(sys)
sys.setdefaultencoding('utf8')

import json,chardet
from Lib.UsualTools import UsualTools

class ResultProcess(object):
    """结果处理
    对DubboService执行完的数据进行结果处理，断言测试结果，生成测试报告，可增加发送邮件或者发送短信等功能。
    """
    def generate_excel_test_report(self,list_with_assertion = [],filepath = "../reports/TestReport.xlsx", step = 0, thread_num = 0):

        """输出excel测试报告
        接口参数为传进来的完整的excellist，和测试报告要输出的路径，"""

        #获取路径名称和文件名称
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        border = Border(left=left, right=right, top=top, bottom=bottom)

        title_font = Font(bold=True,size='13',color=colors.BLUE)

        fill = PatternFill(start_color=colors.BLUE, end_color=colors.BLUE)

        al = Alignment(horizontal="center", vertical="center")

        return_value = self.statistics(list_with_assertion)
        list = filepath.split("/")
        filename_len = len(list[len(list) - 1])
        dirname_len = len(filepath) - filename_len
        filename = list[len(list) - 1]
        dirname = filepath[0: dirname_len - 1]

        logging.info( UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report:  目录名称为: %s".encode('utf8') % dirname)
        logging.info( UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report:  文件名称为: %s".encode('utf8') % filename)

        #判断有没有目录下的这个文件
        if not os.path.isfile(filepath):
            #如果没有这个文件，就判断有没有这个目录
            if not os.path.isdir(dirname):
                #没有这个目录，创建目录
                os.makedirs(dirname)
            #创建文件
            # 新建一个workbook
            wb = Workbook()
            # 第一个sheet是ws

        else:
            wb = Workbook()



        #设置表头
        list = [u'用例编号',u'名称',u'描述',u'系统',u'服务',u'接口',u'数据初始化',u'数据恢复',u'参数',u'预期结果',u'实际结果',u'断言结果',u'测试结果',u'执行时间']
        # 新建一个workbook


        # 第一个sheet是ws
        ws = wb.worksheets[0]
        ws.print_options.horizontalCentered = True
        ws.title = u"interface test"
        ws.column_dimensions["A"].width = 10.0
        ws.column_dimensions["B"].width = 16.0
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 13.0
        ws.column_dimensions["E"].width = 25.0
        ws.column_dimensions["F"].width = 15.0
        ws.column_dimensions["G"].width = 12.0
        ws.column_dimensions["H"].width = 12.0
        ws.column_dimensions["I"].width = 20.0
        ws.column_dimensions["J"].width = 20.0
        ws.column_dimensions["K"].width = 20.0
        ws.column_dimensions["L"].width = 20.0
        ws.column_dimensions["M"].width = 10.0
        ws.column_dimensions["N"].width = 10.0

        #写表头
        for i in range(1,len(list)+1):
            ws.cell(row=1, column=i).alignment = al
            ws.cell(row=1, column=i).font = title_font
            # ws.cell(row=1, column=i).fill = fill
            ws.cell(row=1, column=i, value=list[i-1]).border = border
        # num = 0
        # failnum = 0
        # system_list = {}
        # step = 5000
        # len_of_list = len(list_with_assertion)
        # thread_num = len_of_list%step == 0 and len_of_list/step or len_of_list/step+1
        # last_num = len_of_list % step
        #
        #
        # start_line = 2
        #
        # for i in range(0,thread_num):
        #
        #     end_line = (i == thread_num - 1 and last_num != 0) and start_line+last_num   or start_line + step
        #     print "ResultProcess.py:  generate_excel_test_report:  INTO last_num%s" % last_num
        #     print "ResultProcess.py:  generate_excel_test_report:  INTO LOOP WRITE"
        #     t1 = threading.Thread(target=self.generate_report, args=(ws,start_line,end_line,list_with_assertion))
        #     t1.start()
        #     t1.join()
        #     start_line += step

        self.generate_report(ws,2,len(list_with_assertion)+2,list_with_assertion,border)
        fail_num = 0
        pass_num = 0
        error_num = 0
        for i in range(len(list_with_assertion)):
            if list_with_assertion[i].get('test_result') == 'FAIL':
                fail_num += 1
            elif list_with_assertion[i].get('test_result') == 'PASS':
                pass_num += 1
            else:
                error_num += 1
        if pass_num != 0:
            prints = u'共执行%s条,成功%s条,失败%s条,通过率为%.2f%%' % (pass_num + fail_num, pass_num, fail_num, 100*float(pass_num)/float(pass_num+fail_num))
        else:
            prints = u'共执行%s条,成功%s条,失败%s条,通过率为%s%%' % (pass_num + fail_num, pass_num, fail_num,"0.00")

        logging.info( UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report:  %s"%prints)

        ws.merge_cells(start_row=len(list_with_assertion)+2, start_column=1,end_row=len(list_with_assertion)+2, end_column=14)
        ws.merge_cells(start_row=len(list_with_assertion)+4, start_column=1,end_row=len(list_with_assertion)+4+return_value.count('\n'), end_column=14)

        logging.info( UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report:  %s" % return_value.count('\n'))

        ws.cell(row=len(list_with_assertion)+2, column=1, value=prints)
        # print u"ResultProcess.py:  generate_excel_test_report: %s" % return_value
        vertAlign = Font(vertAlign='subscript')
        ws.cell(row=len(list_with_assertion) + 4, column=1, value=return_value)
        wb.save(filepath)

    def generate_report(self,ws,start_row=0,end_row=0,list_with_assertion=[],border=Border):

        """将结果写入excel"""

        logging.info( UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report: 写入excel最后的行数为%s"%end_row)

        pass_font = Font(color='458B00',bold=True)
        fail_font = Font(color='8B2323',bold=True)
        error_font = Font(color='8B5A00',bold=True)
        for i in range(start_row, end_row):
            j=i-2
            tmpCase = {}
            ws.cell(row=i, column=1).border = border
            tmpCase['case_id'] = ws.cell(row=i, column=1,value=list_with_assertion[j].get('case_id')).font = Font(bold=True)
            tmpCase['name'] = ws.cell(row=i, column=2,value=list_with_assertion[j].get('name')).border = border
            tmpCase['desc'] = ws.cell(row=i, column=3,value=list_with_assertion[j].get('desc')).border = border
            tmpCase['system'] = ws.cell(row=i, column=4,value=list_with_assertion[j].get('system')).border = border
            tmpCase['service'] = ws.cell(row=i, column=5,value=list_with_assertion[j].get('service')).border = border
            tmpCase['method'] = ws.cell(row=i, column=6,value=list_with_assertion[j].get('method')).border = border
            tmpCase['data_init'] = ws.cell(row=i, column=7,value=list_with_assertion[j].get('data_init')).border = border
            tmpCase['data_recover'] = ws.cell(row=i, column=8,value=list_with_assertion[j].get('data_recover')).border = border
            tmpCase['params'] = ws.cell(row=i, column=9,value=list_with_assertion[j].get('params')).border = border
            tmpCase['expect'] = ws.cell(row=i, column=10,value=list_with_assertion[j].get('expect')).border = border
            tmpCase['return_msg'] = ws.cell(row=i, column=11, value=list_with_assertion[j].get('return_msg')).border = border
            tmpCase['assert_msg'] = ws.cell(row=i, column=12, value=list_with_assertion[j].get('assert_msg')).border = border
            tmpCase['test_result'] = ws.cell(row=i, column=13).border = border
            if list_with_assertion[j].get('test_result') == "PASS":
                ws.cell(row=i, column=13, value=list_with_assertion[j].get('test_result')).font = pass_font
            elif list_with_assertion[j].get('test_result') == "FAIL":
                ws.cell(row=i, column=13,value=list_with_assertion[j].get('test_result')).font = fail_font
            else:
                tmpCase['test_result'] = ws.cell(row=i, column=13, value=list_with_assertion[j].get('test_result')).font = error_font
            tmpCase['perform_time'] = ws.cell(row=i, column=14, value=list_with_assertion[j].get('perform_time')).border = border





    def statistics(self,result_list = []):

        """结果统计的方法"""

        system_dicts= {}
        # print '*************'
        # print result_list
        for i in range(len(result_list)):
            if result_list[i]['return_msg'] in 'CASE_ERROR':
                continue

            try:
                # print 'ResultProcess.py:  statistics:  ' % system_dicts[result_list[i]['system']]
                system_dict = system_dicts[result_list[i]['system']]
            except KeyError, e:
                # print "ResultProcess.py:  statistics:  sys KeyError key null"
                system_dict =  system_dicts[result_list[i]['system']] = {}
                system_dict['fail'] = 0
                system_dict['pass'] = 0
                system_dict['count'] = 0
                system_dict['service_count'] = 0
                system_dict['method_count'] = 0

            try:
                # print 'ResultProcess.py:  statistics:  ' %  system_dicts[result_list[i]['system']][result_list[i]['service']]
                service_dict = system_dicts[result_list[i]['system']][result_list[i]['service']]
            except KeyError, e:
                # print "ResultProcess.py:  statistics:  service KeyError key null"
                service_dict = system_dicts[result_list[i]['system']][result_list[i]['service']] = {}
                service_dict['fail'] = 0
                service_dict['pass'] = 0
                service_dict['count'] = 0
                service_dict['method_count'] = 0
                system_dict['service_count'] += 1



            try:

                method_dict = system_dicts[result_list[i]['system']][result_list[i]['service']][result_list[i]['method']]
            except KeyError, e:
                # print "ResultProcess.py:  statistics:  method KeyError key null"

                method_dict = system_dicts[result_list[i]['system']][result_list[i]['service']][result_list[i]['method']] = {}
                method_dict['fail'] = 0
                method_dict['pass'] = 0
                method_dict['count'] = 0
                service_dict['method_count'] += 1
                system_dict['method_count'] += 1

            if result_list[i]['test_result'] == 'PASS':

                method_dict['pass'] += 1
                method_dict['count'] += 1
                # system_list[tmpCase['system'].value][tmpCase['service'].value]['fail'] = 0
                service_dict['pass'] += 1
                service_dict['count'] += 1
                # system_list[tmpCase['system'].value]['fail'] = 0
                system_dict['pass'] += 1
                system_dict['count'] += 1
                # system_dict['method_count'] += 1


            elif result_list[i]['test_result'] == 'FAIL':
                method_dict['fail'] += 1
                method_dict['count'] += 1
                # system_list[tmpCase['system'].value][tmpCase['service'].value]['fail'] = 0
                service_dict['fail'] += 1
                service_dict['count'] += 1
                # system_list[tmpCase['system'].value]['fail'] = 0
                system_dict['fail'] += 1
                system_dict['count'] += 1
                # system_dict['method_count'] += 1
        test_num = ''
        for k, v in system_dicts.items():
            if type(v) != type({}):
                # print "K:v : %s , %d" % (k, v)
                continue
            if system_dicts[k]['pass'] != 0 :
                test_num += u"|❥系统%s中有服务%d组,接口%d个,用例%d条,通过%d条,失败%d条,通过率为%.2f%%\n" % (k, system_dicts[k]['service_count'], system_dicts[k]['method_count'], system_dicts[k]['count'],system_dicts[k]['pass'], system_dicts[k]['fail'],100 * float(system_dicts[k]['pass']) / float(system_dicts[k]['count']))
            else:
                test_num += u"|❥系统%s中有服务%d组,接口%d个,用例%d条,通过%d条,失败%d条,通过率为%s%%\n" % (k,system_dicts[k]['service_count'], system_dicts[k]['method_count'],system_dicts[k]['count'],system_dicts[k]['pass'],system_dicts[k]['fail'],u"0.00")

            # print "system %s in the %d case" %(k,system_list[k]['count'])
            for k2, v2 in system_dicts[k].items():
                if type(v2) != type({}):
                    # print "K2:v2 : %s , %d" % (k2, v2)
                    continue

                if  system_dicts[k][k2]['pass'] != 0 :
                    test_num += u"|---❁服务%s中有接口%d个,用例%s条,成功%d条,失败%d条,通过率为%.2f%%\n" % (k2,system_dicts[k][k2]['method_count'], system_dicts[k][k2]['count'],system_dicts[k][k2]['pass'],system_dicts[k][k2]['fail'],100*float(system_dicts[k][k2]['pass'])/float(system_dicts[k][k2]['count']))
                else:
                    test_num += u"|---❁服务%s中有接口%d个,用例%s条,成功%d条,失败%d条,通过率为%s%%\n" % (k2,system_dicts[k][k2]['method_count'], system_dicts[k][k2]['count'],system_dicts[k][k2]['pass'],system_dicts[k][k2]['fail'],u"0.00")

                # test_num+=( "class: %s, methods: %d" % (k2, len(v2)-3))+'\n'
                for k3, v3 in system_dicts[k][k2].items():
                    if type(v3) != type({}):
                        # print "K3:v3: %s, %d" % (k3, v3)
                        continue

                    # print "%s" % system_list[k][k2][k3]
                    if system_dicts[k][k2][k3]['pass'] != 0 :
                        test_num += u"|------✧接口%s中有用例%d条,成功%d条,失败%d条,通过率为%.2f%%\n" % (k3, system_dicts[k][k2][k3]['count'],system_dicts[k][k2][k3]['pass'],system_dicts[k][k2][k3]['fail'],100*float(system_dicts[k][k2][k3]['pass'])/float(system_dicts[k][k2][k3]['count']))
                    else:
                        test_num += u"|------✧接口%s中有用例%d条,成功%d条,失败%d条,通过率为%s%%\n" % (k3, system_dicts[k][k2][k3]['count'], system_dicts[k][k2][k3]['pass'],system_dicts[k][k2][k3]['fail'],u"0.00")

        logging.info( UsualTools().get_current_time() + u"=>ResultProcess.py: statistics测试统计为 \n%s" % test_num)

        return test_num

    def send_mail(self, receiver, title, content, file):
        """

        :param receiver:
        :param title:
        :param content:
        :param file:
        :return:
        """

        pass


    # def generate_html_test_report(self,list_with_assertion = [],file_path='../报告/index.html',):
    #     """输出html格式的测试报告"""
    #
    #     list = file_path.split("/")
    #     filename_len = len(list[len(list) - 1])
    #     dirname_len = len(file_path) - filename_len
    #     filename = list[len(list) - 1]
    #     dirname = file_path[0: dirname_len - 1]
    #
    #     print UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report:  目录名称为: %s" % dirname
    #     print UsualTools().get_current_time() + u"=>ResultProcess.py:  generate_excel_test_report:  文件名称为: %s" % filename
    #     # 判断有没有目录下的这个文件
    #     if not os.path.isfile(file_path):
    #         # 如果没有这个文件，就判断有没有这个目录
    #         if not os.path.isdir(dirname):
    #             # 没有这个目录，创建目录
    #             os.makedirs(dirname)
    #
    #     return_value = self.statistics(list_with_assertion)
    #     return_value_num = return_value.split('\n')
    #     print "***********"
    #     print len(return_value_num)
    #     html = open(file_path, 'w')
    #     html.write("""
    #     <html>
    #     <head>
    #         <meta http-equiv="content-type" content="text/html;charset=utf-8">
    #       <title>Test</title>
    #       <style>img{float:left;margin:5px;}</style>
    #     </head>
    #     <body>
    #     """)
    #     files = os.listdir('.')
    #     for i in range(0,len(return_value_num)):
    #         html.write("<p>%s</p>" % return_value_num[i])
    #         print '-------------------'
    #         print return_value_num[i]
    #
    #     html.write('</body></html>')
    #     html.close()
    #
    #     page = PyH('My wonderful PyH page')
    #     # page.addCSS('myStylesheet1.css', 'myStylesheet2.css')
    #     # page.addJS('myJavascript1.js', 'myJavascript2.js')
    #     # page.h1('My big title', cl='center')
    #     # page.div(cl='myCSSclass1 myCSSclass2', id='myDiv1').p('I love PyH!', id='myP1')
    #     # mydiv2 = page.div(id='myDiv2')
    #     # mydiv2.h2('A smaller title').p('Followed by a paragraph.')
    #     # page.div(id='myDiv3')
    #     # page.myDiv3.attributes['cl'] = 'myCSSclass3'
    #     # page.myDiv3.p('Another paragraph')
    #     page.printOut(file='../report/index.html')

if __name__ == "__main__":
    case_list = [{'name': '\xe4\xba\xa4\xe6\x98\x93\xe5\x8d\xa0\xe5\xba\x93\xe5\xad\x98\xe6\x8e\xa5\xe5\x8f\xa3',
                  'service': 'StockServices', 'system': 'com.pzj.core.stock.service',
                  'return_msg': '{"errorCode":10002,"data":true,"errorMsg":"ok","ok":true}\r\nelapsed: 11 ms.\r\ndubbo>',
                  'data_recover': 'default', 'case_id': 'TC_002',
                  'params': '{  "transactionId": "abc",  "productId": 2216619736567387,  "stockId": 233,  "stockNum": 1,  "userId": 2216619736563723,  "invokeOnlyId": "asd0810slfdsdadda"}',
                  'expect': '{  "errorCode":10002,  "data":true,  "errorMsg":"ok",  "ok":true}', 'data_init': 'default',
                  'method': 'occupyStock1',
                  'desc': '\xe5\xaf\xb9stock\xe7\x9a\x84occupyStock\xe6\x8e\xa5\xe5\x8f\xa3\xe8\xbf\x9b\xe8\xa1\x8c\xe6\xad\xa3\xe5\xb8\xb8\xe5\x8f\x82\xe6\x95\xb0\xe8\xaf\xb7\xe6\xb1\x82\xe3\x80\x82'},{'name': '\xe4\xba\xa4\xe6\x98\x93\xe5\x8d\xa0\xe5\xba\x93\xe5\xad\x98\xe6\x8e\xa5\xe5\x8f\xa3',
                  'service': 'StockService', 'system': 'com.pzj.core.stock.service',
                  'return_msg': '{"errorCode":10002,"data":true,"errorMsg":"ok","ok":true}\r\nelapsed: 11 ms.\r\ndubbo>',
                  'data_recover': 'default', 'case_id': 'TC_002',
                  'params': '{  "transactionId": "abc",  "productId": 2216619736567387,  "stockId": 233,  "stockNum": 1,  "userId": 2216619736563723,  "invokeOnlyId": "asd0810slfdsdadda"}',
                  'expect': '{  "errorCode":10001,  "data":true,  "errorMsg":"ok",  "ok":true}', 'data_init': 'default',
                  'method': 'occupyStock2',
                  'desc': '\xe5\xaf\xb9stock\xe7\x9a\x84occupyStock\xe6\x8e\xa5\xe5\x8f\xa3\xe8\xbf\x9b\xe8\xa1\x8c\xe6\xad\xa3\xe5\xb8\xb8\xe5\x8f\x82\xe6\x95\xb0\xe8\xaf\xb7\xe6\xb1\x82\xe3\x80\x82'},{'name': '\xe4\xba\xa4\xe6\x98\x93\xe5\x8d\xa0\xe5\xba\x93\xe5\xad\x98\xe6\x8e\xa5\xe5\x8f\xa3',
                  'service': 'StockServices', 'system': 'com.pzj.core.stock.service',
                  'return_msg': 'CASE_ERROR:{"errorCode":10001,"data":true,"errorMsg":"ok","ok":true}\r\nelapsed: 11 ms.\r\ndubbo>',
                  'data_recover': 'default', 'case_id': 'TC_002',
                  'params': '{  "transactionId": "abc",  "productId": 2216619736567387,  "stockId": 233,  "stockNum": 1,  "userId": 2216619736563723,  "invokeOnlyId": "asd0810slfdsdadda"}',
                  'expect': '{  "errorCode":10002,  "data":true,  "errorMsg":"ok",  "ok":true}', 'data_init': 'default',
                  'method': 'occupyStock1',
                  'desc': '\xe5\xaf\xb9stock\xe7\x9a\x84occupyStock\xe6\x8e\xa5\xe5\x8f\xa3\xe8\xbf\x9b\xe8\xa1\x8c\xe6\xad\xa3\xe5\xb8\xb8\xe5\x8f\x82\xe6\x95\xb0\xe8\xaf\xb7\xe6\xb1\x82\xe3\x80\x82'}]

    ResultProcess().generate_excel_test_report(ResultProcess().process_result(case_list))
    # ResultProcess().generate_html_test_report(ResultProcess().process_result(case_list));
